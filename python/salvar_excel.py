from openpyxl import Workbook, load_workbook

CAMPOS = ["hostname","ip","usuario","ram","ssd","status","sistema","ultimo_acesso","alerta","notas"]
arquivo = "inventario.xlsx"

try:
    wb = load_workbook(arquivo)
    ws = wb.active
except:
    wb = Workbook()
    ws = wb.active
    ws.title = "Computadores"
    ws.append([c.capitalize() for c in CAMPOS])
    wb.save(arquivo)
    ws = wb.active

pc = {}
for c in CAMPOS:
    pc[c] = input(f"{c}: ")

ws.append([pc[c] for c in CAMPOS])
wb.save(arquivo)
print(f"{pc['hostname']} salvo na planilha!")